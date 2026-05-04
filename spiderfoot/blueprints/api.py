# -*- coding: utf-8 -*-
# -----------------------------------------------------------------
# Name:         api
# Purpose:      Flask API blueprint — faithful port of sfwebui.py
#               CherryPy endpoints so sfcli.py and external callers
#               keep working with identical JSON shapes.
# -----------------------------------------------------------------

import csv
import html
import json
import logging
import multiprocessing as mp
import re
import sqlite3
import string
import time
import uuid
from copy import deepcopy
from io import BytesIO, StringIO
from operator import itemgetter

from flask import Blueprint, Response, current_app, jsonify, request

import openpyxl

from spiderfoot import SpiderFootDb, SpiderFootHelpers, __version__
from spiderfoot.services.preset_service import serialize_preset, validate_module_names

api_bp = Blueprint('api', __name__)

log = logging.getLogger(f"spiderfoot.{__name__}")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_config():
    """Return the live SF config dict stored on the app."""
    return current_app.config['SF_CONFIG']


def get_default_config():
    """Return the original (default) SF config snapshot."""
    return current_app.config.get('SF_DEFAULT_CONFIG', get_config())


def get_db():
    """Create a SpiderFootDb handle using the current config."""
    return SpiderFootDb(get_config())


def get_logging_queue():
    """Return the multiprocessing logging queue."""
    return current_app.config.get('SF_LOGGING_QUEUE')


def jsonify_error(status, message):
    """Return a JSON error response matching the CherryPy format."""
    resp = jsonify({'error': {'http_status': status, 'message': message}})
    resp.status_code = int(status)
    return resp


def clean_user_input(input_list):
    """Sanitise user input (mirrors SpiderFootWebUi.cleanUserInput)."""
    ret = []
    for item in input_list:
        if not item:
            ret.append('')
            continue
        c = html.escape(item, True)
        ret.append(c)
    return ret


def _safe_filename(name: str) -> str:
    """Sanitize a string for use in Content-Disposition filename."""
    if not name:
        return "SpiderFoot"
    return re.sub(r'[^\w\-.]', '_', name)[:100]


_SENSITIVE_OPT_PATTERNS = ('api_key', 'apikey', 'password', 'secret', 'token', 'passphrase')


def _csv_safe(value) -> str:
    """Prevent CSV formula injection by prefixing dangerous characters."""
    s = str(value) if value is not None else ''
    if s and s[0] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + s
    return s


def search_base(id=None, eventType=None, value=None):
    """Core search logic (mirrors SpiderFootWebUi.searchBase)."""
    retdata = []

    if not id and not eventType and not value:
        return retdata

    if not value:
        value = ''

    regex = ""
    if value.startswith("/") and value.endswith("/"):
        regex = value[1:len(value) - 1]
        value = ""

    value = value.replace('*', '%')
    if value in [None, ""] and regex in [None, ""]:
        value = "%"
        regex = ""

    dbh = get_db()
    criteria = {
        'scan_id': id or '',
        'type': eventType or '',
        'value': value or '',
        'regex': regex or '',
    }

    try:
        data = dbh.search(criteria)
    except Exception:
        return retdata

    for row in data:
        lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
        escapeddata = html.escape(row[1])
        escapedsrc = html.escape(row[2])
        retdata.append([lastseen, escapeddata, escapedsrc,
                        row[3], row[5], row[6], row[7], row[8], row[10],
                        row[11], row[4], row[13], row[14]])

    return retdata


def build_excel(data, column_names, sheet_name_index=0):
    """Build an Excel workbook from data rows (mirrors buildExcel)."""
    row_nums = dict()
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    column_names.pop(sheet_name_index)
    allowed_sheet_chars = string.ascii_uppercase + string.digits + '_'

    for row in data:
        sheet_name = "".join(
            [c for c in str(row.pop(sheet_name_index)) if c.upper() in allowed_sheet_chars]
        )
        try:
            sheet = workbook[sheet_name]
        except KeyError:
            workbook.create_sheet(sheet_name)
            sheet = workbook[sheet_name]
            for col_num, column_title in enumerate(column_names, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = column_title
            row_nums[sheet_name] = 2

        for col_num, cell_value in enumerate(row, 1):
            cell = sheet.cell(row=row_nums[sheet_name], column=col_num)
            cell.value = cell_value

        row_nums[sheet_name] += 1

    if row_nums:
        workbook.remove(default_sheet)

    workbook._sheets.sort(key=lambda ws: ws.title)

    with BytesIO() as f:
        workbook.save(f)
        f.seek(0)
        return f.read()


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@api_bp.route('/ping')
def ping():
    """CLI connectivity check."""
    return jsonify(["SUCCESS", __version__])


@api_bp.route('/query', methods=['POST'])
def query():
    """Run a SELECT query against the database (CLI).

    Uses a read-only SQLite connection. Rejects semicolons and non-SELECTs.
    """
    import sqlite3

    q = request.values.get('query', '').strip()

    if not q:
        return jsonify_error('400', "Invalid query.")

    if not q.lower().startswith("select"):
        return jsonify_error('400', "Only SELECT queries are allowed.")

    if ';' in q:
        return jsonify_error('400', "Semicolons are not allowed in queries.")

    q_lower = q.lower()
    for keyword in ('attach', 'pragma', 'load_extension'):
        if re.search(r'\b' + keyword + r'\b', q_lower):
            return jsonify_error('400', f"'{keyword}' is not permitted.")
    for table in ('tbl_config',):
        if table in q_lower:
            return jsonify_error('400', f"Access to {table} is not permitted via this endpoint.")

    # Use the same database path as the main connection
    db_path = current_app.config.get('SF_CONFIG', {}).get('__database', '')
    if not db_path:
        return jsonify_error('500', "Database not configured.")

    try:
        conn = sqlite3.connect(f'file:{db_path}?mode=ro', uri=True)
        try:
            cursor = conn.execute(q)
            data = cursor.fetchall()
            column_names = [c[0] for c in cursor.description]
            return jsonify([dict(zip(column_names, row)) for row in data])
        finally:
            conn.close()
    except Exception:
        current_app.logger.warning(f"Query endpoint error for query: {q}", exc_info=True)
        return jsonify_error('500', "Query failed.")


# -- Scan list ---------------------------------------------------------------

@api_bp.route('/scanlist')
def scanlist():
    """List all scans."""
    dbh = get_db()
    data = dbh.scanInstanceList()
    retdata = []

    for row in data:
        created = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[3]))
        riskmatrix = {"HIGH": 0, "MEDIUM": 0, "LOW": 0, "INFO": 0}
        try:
            correlations = dbh.scanCorrelationSummary(row[0], by="risk")
            if correlations:
                for c in correlations:
                    riskmatrix[c[0]] = c[1]
        except Exception:
            pass

        if row[4] == 0:
            started = "Not yet"
        else:
            started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[4]))

        if row[5] == 0:
            finished = "Not yet"
        else:
            finished = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[5]))

        retdata.append([row[0], row[1], row[2], created, started, finished, row[6], row[7], riskmatrix])

    return jsonify(retdata)


# -- Scan status -------------------------------------------------------------

@api_bp.route('/scanstatus', methods=['GET', 'POST'])
def scanstatus():
    """Basic scan info including status and risk matrix."""
    id = request.values.get('id', '')
    dbh = get_db()
    data = dbh.scanInstanceGet(id)

    if not data:
        return jsonify([])

    created = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[2]))
    started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[3]))
    ended = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[4]))
    riskmatrix = {"HIGH": 0, "MEDIUM": 0, "LOW": 0, "INFO": 0}
    try:
        correlations = dbh.scanCorrelationSummary(id, by="risk")
        if correlations:
            for c in correlations:
                riskmatrix[c[0]] = c[1]
    except Exception:
        pass

    return jsonify([data[0], data[1], created, started, ended, data[5], riskmatrix])


# -- Scan summary ------------------------------------------------------------

@api_bp.route('/scansummary', methods=['GET', 'POST'])
def scansummary():
    """Summary of scan results."""
    id = request.values.get('id', '')
    by = request.values.get('by', '')
    retdata = []

    dbh = get_db()

    try:
        scandata = dbh.scanResultSummary(id, by)
    except Exception:
        return jsonify(retdata)

    try:
        statusdata = dbh.scanInstanceGet(id)
    except Exception:
        return jsonify(retdata)

    for row in scandata:
        if row[0] == "ROOT":
            continue
        lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[2]))
        retdata.append([row[0], row[1], lastseen, row[3], row[4], statusdata[5]])

    return jsonify(retdata)


# -- Scan correlations -------------------------------------------------------

@api_bp.route('/scancorrelations', methods=['GET', 'POST'])
def scancorrelations():
    """Correlation results from a scan."""
    id = request.values.get('id', '')
    retdata = []

    dbh = get_db()
    try:
        corrdata = dbh.scanCorrelationList(id)
    except Exception:
        return jsonify(retdata)

    for row in corrdata:
        retdata.append([row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]])

    return jsonify(retdata)


# -- Scan event results ------------------------------------------------------

@api_bp.route('/scaneventresults', methods=['GET', 'POST'])
def scaneventresults():
    """All event results for a scan."""
    id = request.values.get('id', '')
    eventType = request.values.get('eventType', None)
    filterfp = request.values.get('filterfp', False)
    correlationId = request.values.get('correlationId', None)
    retdata = []

    dbh = get_db()

    if not eventType:
        eventType = 'ALL'

    try:
        data = dbh.scanResultEvent(id, eventType, filterfp, correlationId=correlationId)
    except Exception:
        return jsonify(retdata)

    for row in data:
        lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
        retdata.append([
            lastseen,
            html.escape(row[1]),
            html.escape(row[2]),
            row[3],
            row[5],
            row[6],
            row[7],
            row[8],
            row[13],
            row[14],
            row[4]
        ])

    return jsonify(retdata)


# -- Scan event results unique -----------------------------------------------

@api_bp.route('/scaneventresultsunique', methods=['GET', 'POST'])
def scaneventresultsunique():
    """Unique event results for a scan."""
    id = request.values.get('id', '')
    eventType = request.values.get('eventType', '')
    filterfp = request.values.get('filterfp', False)

    dbh = get_db()
    retdata = []

    try:
        data = dbh.scanResultEventUnique(id, eventType, filterfp)
    except Exception:
        return jsonify(retdata)

    for row in data:
        escaped = html.escape(row[0])
        retdata.append([escaped, row[1], row[2]])

    return jsonify(retdata)


# -- Scan options ------------------------------------------------------------

@api_bp.route('/scanopts', methods=['GET', 'POST'])
def scanopts():
    """Configuration used for the specified scan."""
    id = request.values.get('id', '')
    config = get_config()
    dbh = get_db()
    ret = dict()

    meta = dbh.scanInstanceGet(id)
    if not meta:
        return jsonify(ret)

    if meta[3] != 0:
        started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(meta[3]))
    else:
        started = "Not yet"

    if meta[4] != 0:
        finished = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(meta[4]))
    else:
        finished = "Not yet"

    ret['meta'] = [meta[0], meta[1], meta[2], started, finished, meta[5]]
    ret['config'] = dbh.scanConfigGet(id)
    ret['configdesc'] = dict()
    for key in list(ret['config'].keys()):
        if ':' not in key:
            globaloptdescs = config.get('__globaloptdescs__')
            if globaloptdescs:
                ret['configdesc'][key] = globaloptdescs.get(key, f"{key} (legacy)")
        else:
            [modName, modOpt] = key.split(':')
            if modName not in list(config.get('__modules__', {}).keys()):
                continue
            if modOpt not in list(config['__modules__'][modName].get('optdescs', {}).keys()):
                continue
            ret['configdesc'][key] = config['__modules__'][modName]['optdescs'][modOpt]

    return jsonify(ret)


# -- Scan log ----------------------------------------------------------------

@api_bp.route('/scanlog', methods=['GET', 'POST'])
def scanlog():
    """Scan log data."""
    id = request.values.get('id', '')
    limit = request.values.get('limit', None)
    rowId = request.values.get('rowId', None)
    reverse = request.values.get('reverse', None)

    dbh = get_db()
    retdata = []

    try:
        data = dbh.scanLogs(id, limit, rowId, reverse)
    except Exception:
        return jsonify(retdata)

    for row in data:
        generated = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000))
        retdata.append([generated, row[1], row[2], html.escape(row[3]), row[4]])

    return jsonify(retdata)


# -- Scan errors -------------------------------------------------------------

@api_bp.route('/scanerrors', methods=['GET', 'POST'])
def scanerrors():
    """Scan error data."""
    id = request.values.get('id', '')
    limit = request.values.get('limit', None)

    dbh = get_db()
    retdata = []

    try:
        data = dbh.scanErrors(id, limit)
    except Exception:
        return jsonify(retdata)

    for row in data:
        generated = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000))
        retdata.append([generated, row[1], html.escape(str(row[2]))])

    return jsonify(retdata)


# -- Scan delete -------------------------------------------------------------

@api_bp.route('/scandelete', methods=['POST'])
def scandelete():
    """Delete scan(s)."""
    id = request.values.get('id', '')

    if not id:
        return jsonify_error('404', "No scan specified")

    dbh = get_db()
    ids = id.split(',')

    for scan_id in ids:
        res = dbh.scanInstanceGet(scan_id)
        if not res:
            return jsonify_error('404', f"Scan {scan_id} does not exist")
        if res[5] in ["RUNNING", "STARTING", "STARTED"]:
            return jsonify_error('400', f"Scan {scan_id} is {res[5]}. You cannot delete running scans.")

    for scan_id in ids:
        dbh.scanInstanceDelete(scan_id)

    return jsonify("")


# -- Stop scan ---------------------------------------------------------------

@api_bp.route('/stopscan', methods=['POST'])
def stopscan():
    """Stop a running scan."""
    id = request.values.get('id', '')

    if not id:
        return jsonify_error('404', "No scan specified")

    dbh = get_db()
    ids = id.split(',')

    for scan_id in ids:
        res = dbh.scanInstanceGet(scan_id)
        if not res:
            return jsonify_error('404', f"Scan {scan_id} does not exist")

        scan_status = res[5]

        if scan_status == "FINISHED":
            return jsonify_error('400', f"Scan {scan_id} has already finished.")
        if scan_status == "ABORTED":
            return jsonify_error('400', f"Scan {scan_id} has already aborted.")
        if scan_status not in ["RUNNING", "STARTING"]:
            return jsonify_error('400', f"The running scan is currently in the state '{scan_status}', please try again later or restart SpiderFoot.")

    for scan_id in ids:
        dbh.scanInstanceSet(scan_id, status="ABORT-REQUESTED")

    return jsonify("")


# -- Start scan --------------------------------------------------------------

@api_bp.route('/startscan', methods=['POST'])
def startscan():
    """Initiate a new scan."""
    from sflib import SpiderFoot
    from sfscan import startSpiderFootScanner

    scanname = request.values.get('scanname', '')
    scantarget = request.values.get('scantarget', '')
    modulelist = request.values.get('modulelist', '')
    typelist = request.values.get('typelist', '')
    usecase = request.values.get('usecase', '')

    config = get_config()

    # Rate limit: reject if too many scans are already running
    max_scans = config.get('_maxscans', 10)
    dbh = get_db()
    running_scans = dbh.scanInstanceList()
    running_count = sum(
        1 for scan in running_scans
        if scan[5] in ('RUNNING', 'STARTED', 'STARTING')
    )
    if running_count >= max_scans:
        return jsonify_error(
            '429',
            f"Maximum concurrent scans ({max_scans}) reached. "
            f"Please wait for a running scan to complete."
        )

    scanname = clean_user_input([scanname])[0]
    scantarget = clean_user_input([scantarget])[0]

    if not scanname:
        return jsonify(["ERROR", "Incorrect usage: scan name was not specified."])

    if not scantarget:
        return jsonify(["ERROR", "Incorrect usage: scan target was not specified."])

    if not typelist and not modulelist and not usecase:
        return jsonify(["ERROR", "Incorrect usage: no modules specified for scan."])

    targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
    if targetType is None:
        return jsonify(["ERROR", "Unrecognised target type."])

    dbh = get_db()
    cfg = deepcopy(config)
    sf = SpiderFoot(cfg)

    modlist = list()

    # User selected modules
    if modulelist:
        modlist = modulelist.replace('module_', '').split(',')

    # User selected types
    if len(modlist) == 0 and typelist:
        typesx = typelist.replace('type_', '').split(',')
        modlist = sf.modulesProducing(typesx)
        newmods = deepcopy(modlist)
        newmodcpy = deepcopy(newmods)

        while len(newmodcpy) > 0:
            for etype in sf.eventsToModules(newmodcpy):
                xmods = sf.modulesProducing([etype])
                for mod in xmods:
                    if mod not in modlist:
                        modlist.append(mod)
                        newmods.append(mod)
            newmodcpy = deepcopy(newmods)
            newmods = list()

    # User selected a use case
    if len(modlist) == 0 and usecase:
        for mod in config.get('__modules__', {}):
            if usecase == 'all' or usecase in config['__modules__'][mod].get('group', []):
                modlist.append(mod)

    if not modlist:
        return jsonify(["ERROR", "Incorrect usage: no modules specified for scan."])

    # Add mandatory storage module
    if "sfp__stor_db" not in modlist:
        modlist.append("sfp__stor_db")
    modlist.sort()

    # Remove stdout module
    if "sfp__stor_stdout" in modlist:
        modlist.remove("sfp__stor_stdout")

    # For private IPs, strip modules that only work with public IPs
    if targetType in ('IP_ADDRESS', 'IPV6_ADDRESS', 'NETBLOCK_OWNER', 'NETBLOCKV6_OWNER'):
        if SpiderFootHelpers.isPrivateIP(scantarget):
            allowed = SpiderFootHelpers.PRIVATE_IP_COMPATIBLE_MODULES
            modlist = [m for m in modlist if m in allowed]
            if not modlist or modlist == ['sfp__stor_db']:
                return jsonify(["ERROR", "No modules available for private IP targets."])

    if targetType in ["HUMAN_NAME", "USERNAME", "BITCOIN_ADDRESS"]:
        scantarget = scantarget.replace("\"", "")
    else:
        scantarget = scantarget.lower()

    scanId = SpiderFootHelpers.genScanInstanceId()
    logging_queue = get_logging_queue()

    try:
        p = mp.Process(target=startSpiderFootScanner, args=(logging_queue, scanname, scanId, scantarget, targetType, modlist, cfg))
        p.daemon = True
        p.start()
    except Exception as e:
        log.error(f"[-] Scan [{scanId}] failed: {e}")
        return jsonify(["ERROR", f"[-] Scan [{scanId}] failed: {e}"])

    # Wait until the scan has initialized (timeout after 30s)
    wait = 0
    while dbh.scanInstanceGet(scanId) is None:
        log.info("Waiting for the scan to initialize...")
        time.sleep(1)
        wait += 1
        if wait >= 30:
            return jsonify(["ERROR", f"Scan [{scanId}] failed to initialize within 30 seconds."])

    return jsonify(["SUCCESS", scanId])


# -- Rerun scan --------------------------------------------------------------

@api_bp.route('/rerunscan', methods=['POST'])
def rerunscan():
    """Rerun a scan."""
    from sflib import SpiderFoot
    from sfscan import startSpiderFootScanner

    id = request.values.get('id', '')
    config = get_config()
    cfg = deepcopy(config)
    dbh = SpiderFootDb(cfg)
    info = dbh.scanInstanceGet(id)

    if not info:
        return jsonify(["ERROR", "Invalid scan ID."])

    scanname = info[0]
    scantarget = info[1]

    scanconfig = dbh.scanConfigGet(id)
    if not scanconfig:
        return jsonify(["ERROR", f"Error loading config from scan: {id}"])

    modlist = scanconfig['_modulesenabled'].split(',')
    if "sfp__stor_stdout" in modlist:
        modlist.remove("sfp__stor_stdout")

    targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
    if not targetType:
        targetType = SpiderFootHelpers.targetTypeFromString(f'"{scantarget}"')

    if targetType not in ["HUMAN_NAME", "BITCOIN_ADDRESS"]:
        scantarget = scantarget.lower()

    scanId = SpiderFootHelpers.genScanInstanceId()
    logging_queue = get_logging_queue()

    try:
        p = mp.Process(target=startSpiderFootScanner, args=(logging_queue, scanname, scanId, scantarget, targetType, modlist, cfg))
        p.daemon = True
        p.start()
    except Exception as e:
        log.error(f"[-] Scan [{scanId}] failed: {e}")
        return jsonify(["ERROR", f"[-] Scan [{scanId}] failed: {e}"])

    wait = 0
    while dbh.scanInstanceGet(scanId) is None:
        log.info("Waiting for the scan to initialize...")
        time.sleep(1)
        wait += 1
        if wait >= 30:
            return jsonify(["ERROR", f"Scan [{scanId}] failed to initialize within 30 seconds."])

    return jsonify(["SUCCESS", scanId])


# -- Rerun scan multi --------------------------------------------------------

@api_bp.route('/rerunscanmulti', methods=['POST'])
def rerunscanmulti():
    """Rerun multiple scans."""
    from sflib import SpiderFoot
    from sfscan import startSpiderFootScanner

    ids = request.values.get('ids', '')
    config = get_config()
    cfg = deepcopy(config)
    dbh = SpiderFootDb(cfg)
    scan_ids = []

    for id in ids.split(","):
        info = dbh.scanInstanceGet(id)
        if not info:
            return jsonify(["ERROR", "Invalid scan ID."])

        scanconfig = dbh.scanConfigGet(id)
        scanname = info[0]
        scantarget = info[1]

        if len(scanconfig) == 0:
            return jsonify(["ERROR", "Something went wrong internally."])

        modlist = scanconfig['_modulesenabled'].split(',')
        if "sfp__stor_stdout" in modlist:
            modlist.remove("sfp__stor_stdout")

        targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
        if targetType is None:
            return jsonify(["ERROR", "Invalid target type. Could not recognize it as a target SpiderFoot supports."])

        scanId = SpiderFootHelpers.genScanInstanceId()
        logging_queue = get_logging_queue()

        try:
            p = mp.Process(target=startSpiderFootScanner, args=(logging_queue, scanname, scanId, scantarget, targetType, modlist, cfg))
            p.daemon = True
            p.start()
        except Exception as e:
            log.error(f"[-] Scan [{scanId}] failed: {e}")
            return jsonify(["ERROR", f"[-] Scan [{scanId}] failed: {e}"])

        wait = 0
        while dbh.scanInstanceGet(scanId) is None:
            log.info("Waiting for the scan to initialize...")
            time.sleep(1)
            wait += 1
            if wait >= 30:
                return jsonify(["ERROR", f"Scan [{scanId}] failed to initialize within 30 seconds."])

        scan_ids.append(scanId)

    return jsonify(["SUCCESS", scan_ids])


# -- Search ------------------------------------------------------------------

@api_bp.route('/search', methods=['GET', 'POST'])
def search():
    """Search scans."""
    id = request.values.get('id', None)
    eventType = request.values.get('eventType', None)
    value = request.values.get('value', None)

    try:
        return jsonify(search_base(id, eventType, value))
    except Exception:
        return jsonify([])


# -- Scan history ------------------------------------------------------------

@api_bp.route('/scanhistory', methods=['GET', 'POST'])
def scanhistory():
    """Historical data for a scan."""
    id = request.values.get('id', '')

    if not id:
        return jsonify_error('404', "No scan specified")

    dbh = get_db()

    try:
        return jsonify(dbh.scanResultHistory(id))
    except Exception:
        return jsonify([])


# -- Scan element type discovery ---------------------------------------------

@api_bp.route('/scanelementtypediscovery', methods=['GET', 'POST'])
def scanelementtypediscovery():
    """Scan element type discovery."""
    id = request.values.get('id', '')
    eventType = request.values.get('eventType', '')

    dbh = get_db()
    pc = dict()
    datamap = dict()
    retdata = dict()

    try:
        leafSet = dbh.scanResultEvent(id, eventType)
        [datamap, pc] = dbh.scanElementSourcesAll(id, leafSet)
    except Exception:
        return jsonify(retdata)

    pc.pop('ROOT', None)
    retdata['tree'] = SpiderFootHelpers.dataParentChildToTree(pc)
    retdata['data'] = datamap

    return jsonify(retdata)


# -- Event types -------------------------------------------------------------

@api_bp.route('/eventtypes')
def eventtypes():
    """List all event types."""
    dbh = get_db()
    types = dbh.eventTypes()
    ret = []

    for r in types:
        ret.append([r[1], r[0]])

    return jsonify(sorted(ret, key=itemgetter(0)))


# -- Modules -----------------------------------------------------------------

@api_bp.route('/modules')
def modules():
    """List all modules."""
    config = get_config()
    ret = []

    modinfo = list(config.get('__modules__', {}).keys())
    if not modinfo:
        return jsonify(ret)

    modinfo.sort()

    for m in modinfo:
        if "__" in m:
            continue
        ret.append({'name': m, 'descr': config['__modules__'][m]['descr']})

    return jsonify(ret)


# -- Correlation rules -------------------------------------------------------

@api_bp.route('/correlationrules')
def correlationrules():
    """List all correlation rules."""
    config = get_config()
    ret = []

    rules = config.get('__correlationrules__', [])
    if not rules:
        return jsonify(ret)

    for r in rules:
        ret.append({
            'id': r['id'],
            'name': r['meta']['name'],
            'descr': r['meta']['description'],
            'risk': r['meta']['risk'],
        })

    return jsonify(ret)


# -- Options raw -------------------------------------------------------------

@api_bp.route('/optsraw', methods=['GET', 'POST'])
def optsraw():
    """Return global and module settings as JSON."""
    config = get_config()
    ret = dict()

    for opt in config:
        if not opt.startswith('__'):
            ret["global." + opt] = config[opt]
            continue

        if opt == '__modules__':
            for mod in sorted(config['__modules__'].keys()):
                for mo in sorted(config['__modules__'][mod]['opts'].keys()):
                    if mo.startswith("_"):
                        continue
                    ret["module." + mod + "." + mo] = config['__modules__'][mod]['opts'][mo]

    return jsonify(['SUCCESS', {'data': ret}])


# -- Options export ----------------------------------------------------------

@api_bp.route('/optsexport', methods=['GET', 'POST'])
def optsexport():
    """Export configuration."""
    from sflib import SpiderFoot

    config = get_config()
    pattern = request.values.get('pattern', None)

    sf = SpiderFoot(config)
    conf = sf.configSerialize(config)
    content = ""

    for opt in sorted(conf):
        if ":_" in opt or opt.startswith("_"):
            continue
        # Skip sensitive options (API keys, passwords, etc.)
        opt_lower = opt.lower()
        if any(p in opt_lower for p in _SENSITIVE_OPT_PATTERNS):
            continue
        if pattern:
            if pattern in opt:
                content += f"{opt}={conf[opt]}\n"
        else:
            content += f"{opt}={conf[opt]}\n"

    return Response(
        content,
        mimetype='text/plain',
        headers={'Content-Disposition': 'attachment; filename="SpiderFoot.cfg"'}
    )


# -- Save settings raw ------------------------------------------------------

@api_bp.route('/savesettingsraw', methods=['POST'])
def savesettingsraw():
    """Save settings (raw/JSON), also used for reset to default."""
    from sflib import SpiderFoot

    allopts = request.values.get('allopts', '')

    config = get_config()

    # Reset config to default
    if allopts == "RESET":
        try:
            dbh = get_db()
            dbh.configClear()
            default_config = get_default_config()
            current_app.config['SF_CONFIG'] = deepcopy(default_config)
        except Exception:
            return jsonify(["ERROR", "Failed to reset settings"])
        return jsonify(["SUCCESS", ""])

    # Save settings
    try:
        dbh = get_db()
        useropts = json.loads(allopts)
        cleanopts = dict()
        for opt in list(useropts.keys()):
            val = clean_user_input([useropts[opt]])[0]
            # Skip masked API key placeholders — preserve existing value
            if val == '********':
                continue
            cleanopts[opt] = val

        currentopts = deepcopy(config)
        sf = SpiderFoot(config)
        new_config = sf.configUnserialize(cleanopts, currentopts)
        current_app.config['SF_CONFIG'] = new_config
        dbh.configSet(sf.configSerialize(new_config))
    except Exception as e:
        return jsonify(["ERROR", f"Processing one or more of your inputs failed: {e}"])

    return jsonify(["SUCCESS", ""])


# -- Save settings (form-based) ---------------------------------------------

@api_bp.route('/savesettings', methods=['POST'])
def savesettings():
    """Save settings via form post (used by web UI)."""
    from sflib import SpiderFoot

    allopts = request.values.get('allopts', '')
    config_file = request.files.get('configFile', None)
    # The frontend import sends file contents as a form field, not a file upload
    config_text = request.values.get('configFile', '')

    config = get_config()

    if config_file or config_text:
        try:
            if config_file:
                contents = config_file.read()
                if isinstance(contents, bytes):
                    contents = contents.decode('utf-8')
            else:
                contents = config_text

            tmp = dict()
            for line in contents.split("\n"):
                if "=" not in line:
                    continue
                opt_array = line.strip().split("=")
                if len(opt_array) == 1:
                    opt_array.append("")
                tmp[opt_array[0]] = '='.join(opt_array[1:])

            allopts = json.dumps(tmp)
        except Exception as e:
            return jsonify(["ERROR", f"Failed to parse input file. Was it generated from SpiderFoot? ({e})"])

    # Reset
    if allopts == "RESET":
        try:
            dbh = get_db()
            dbh.configClear()
            default_config = get_default_config()
            current_app.config['SF_CONFIG'] = deepcopy(default_config)
        except Exception:
            return jsonify(["ERROR", "Failed to reset settings"])
        return jsonify(["SUCCESS", ""])

    try:
        dbh = get_db()
        useropts = json.loads(allopts)
        cleanopts = dict()
        for opt in list(useropts.keys()):
            val = clean_user_input([useropts[opt]])[0]
            # Skip masked API key placeholders — preserve existing value
            if val == '********':
                continue
            cleanopts[opt] = val

        currentopts = deepcopy(config)
        sf = SpiderFoot(config)
        new_config = sf.configUnserialize(cleanopts, currentopts)
        current_app.config['SF_CONFIG'] = new_config
        dbh.configSet(sf.configSerialize(new_config))
    except Exception as e:
        return jsonify(["ERROR", f"Processing one or more of your inputs failed: {e}"])

    return jsonify(["SUCCESS", ""])


# -- Result set false positive -----------------------------------------------

@api_bp.route('/resultsetfp', methods=['POST'])
def resultsetfp():
    """Set results as false positive."""
    id = request.values.get('id', '')
    resultids = request.values.get('resultids', '')
    fp = request.values.get('fp', '')

    dbh = get_db()

    if fp not in ["0", "1"]:
        return jsonify(["ERROR", "No FP flag set or not set correctly."])

    try:
        ids = json.loads(resultids)
    except Exception:
        return jsonify(["ERROR", "No IDs supplied."])

    status = dbh.scanInstanceGet(id)
    if not status:
        return jsonify(["ERROR", f"Invalid scan ID: {id}"])

    if status[5] not in ["ABORTED", "FINISHED", "ERROR-FAILED"]:
        return jsonify(["WARNING", "Scan must be in a finished state when setting False Positives."])

    if fp == "0":
        data = dbh.scanElementSourcesDirect(id, ids)
        for row in data:
            if str(row[14]) == "1":
                return jsonify(["WARNING", f"Cannot unset element {id} as False Positive if a parent element is still False Positive."])

    childs = dbh.scanElementChildrenAll(id, ids)
    allIds = ids + childs

    ret = dbh.scanResultsUpdateFP(id, allIds, fp)
    if ret:
        return jsonify(["SUCCESS", ""])

    return jsonify(["ERROR", "Exception encountered."])


# -- Vacuum ------------------------------------------------------------------

@api_bp.route('/vacuum', methods=['POST'])
def vacuum():
    """Vacuum the database."""
    dbh = get_db()
    try:
        if dbh.vacuumDB():
            return jsonify(["SUCCESS", ""])
        return jsonify(["ERROR", "Vacuuming the database failed"])
    except Exception as e:
        return jsonify(["ERROR", f"Vacuuming the database failed: {e}"])


# -- Scan export logs (CSV) --------------------------------------------------

@api_bp.route('/scanexportlogs', methods=['GET', 'POST'])
def scanexportlogs():
    """Export scan logs as CSV."""
    id = request.values.get('id', '')
    dialect = request.values.get('dialect', 'excel')

    dbh = get_db()

    try:
        data = dbh.scanLogs(id, None, None, True)
    except Exception:
        return jsonify(["ERROR", "Scan ID not found."])

    if not data:
        return jsonify(["ERROR", "Scan ID not found."])

    fileobj = StringIO()
    parser = csv.writer(fileobj, dialect=dialect)
    parser.writerow(["Date", "Component", "Type", "Event", "Event ID"])
    for row in data:
        parser.writerow([
            time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000)),
            _csv_safe(row[1]),
            _csv_safe(row[2]),
            _csv_safe(row[3]),
            _csv_safe(row[4])
        ])

    return Response(
        fileobj.getvalue().encode('utf-8'),
        mimetype='application/csv',
        headers={
            'Content-Disposition': f'attachment; filename="SpiderFoot-{_safe_filename(str(id))}.log.csv"',
            'Pragma': 'no-cache'
        }
    )


# -- Scan correlations export ------------------------------------------------

@api_bp.route('/scancorrelationsexport', methods=['GET', 'POST'])
def scancorrelationsexport():
    """Export scan correlation data as CSV or Excel."""
    id = request.values.get('id', '')
    filetype = request.values.get('filetype', 'csv')
    dialect = request.values.get('dialect', 'excel')

    dbh = get_db()

    try:
        scaninfo = dbh.scanInstanceGet(id)
        scan_name = scaninfo[0]
    except Exception:
        return jsonify(["ERROR", "Could not retrieve info for scan."])

    try:
        correlations = dbh.scanCorrelationList(id)
    except Exception:
        return jsonify(["ERROR", "Could not retrieve correlations for scan."])

    headings = ["Rule Name", "Correlation", "Risk", "Description"]

    if filetype.lower() in ["xlsx", "excel"]:
        rows = []
        for row in correlations:
            correlation = row[1]
            rule_name = row[2]
            rule_risk = row[3]
            rule_description = row[5]
            rows.append([rule_name, correlation, rule_risk, rule_description])

        fname = f"{_safe_filename(scan_name)}-SpiderFoot-correlations.xlsx"

        return Response(
            build_excel(rows, headings, sheet_name_index=0),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{fname}"',
                'Pragma': 'no-cache'
            }
        )

    if filetype.lower() == 'csv':
        fileobj = StringIO()
        parser = csv.writer(fileobj, dialect=dialect)
        parser.writerow(headings)

        for row in correlations:
            correlation = row[1]
            rule_name = row[2]
            rule_risk = row[3]
            rule_description = row[5]
            parser.writerow([_csv_safe(rule_name), _csv_safe(correlation), _csv_safe(rule_risk), _csv_safe(rule_description)])

        fname = f"{_safe_filename(scan_name)}-SpiderFoot-correlations.csv"

        return Response(
            fileobj.getvalue().encode('utf-8'),
            mimetype='application/csv',
            headers={
                'Content-Disposition': f'attachment; filename="{fname}"',
                'Pragma': 'no-cache'
            }
        )

    return jsonify(["ERROR", "Invalid export filetype."])


# -- Scan event result export ------------------------------------------------

@api_bp.route('/scaneventresultexport', methods=['GET', 'POST'])
def scaneventresultexport():
    """Export scan event result data as CSV or Excel."""
    id = request.values.get('id', '')
    type_ = request.values.get('type', '')
    filetype = request.values.get('filetype', 'csv')
    dialect = request.values.get('dialect', 'excel')

    dbh = get_db()
    data = dbh.scanResultEvent(id, type_)

    if filetype.lower() in ["xlsx", "excel"]:
        rows = []
        for row in data:
            if row[4] == "ROOT":
                continue
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
            rows.append([lastseen, str(row[4]), str(row[3]), str(row[2]), row[13], datafield])

        return Response(
            build_excel(rows, ["Updated", "Type", "Module", "Source", "F/P", "Data"], sheet_name_index=1),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': 'attachment; filename=SpiderFoot.xlsx',
                'Pragma': 'no-cache'
            }
        )

    if filetype.lower() == 'csv':
        fileobj = StringIO()
        parser = csv.writer(fileobj, dialect=dialect)
        parser.writerow(["Updated", "Type", "Module", "Source", "F/P", "Data"])
        for row in data:
            if row[4] == "ROOT":
                continue
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
            parser.writerow([lastseen, _csv_safe(row[4]), _csv_safe(row[3]), _csv_safe(row[2]), _csv_safe(row[13]), _csv_safe(datafield)])

        return Response(
            fileobj.getvalue().encode('utf-8'),
            mimetype='application/csv',
            headers={
                'Content-Disposition': 'attachment; filename=SpiderFoot.csv',
                'Pragma': 'no-cache'
            }
        )

    return jsonify(["ERROR", "Invalid export filetype."])


# -- Scan event result export multi ------------------------------------------

@api_bp.route('/scaneventresultexportmulti', methods=['GET', 'POST'])
def scaneventresultexportmulti():
    """Export scan event result data for multiple scans."""
    ids = request.values.get('ids', '')
    filetype = request.values.get('filetype', 'csv')
    dialect = request.values.get('dialect', 'excel')

    dbh = get_db()
    scaninfo = dict()
    data = list()
    scan_name = ""

    for id in ids.split(','):
        scaninfo[id] = dbh.scanInstanceGet(id)
        if scaninfo[id] is None:
            continue
        scan_name = scaninfo[id][0]
        data = data + dbh.scanResultEvent(id)

    if not data:
        return Response('', status=204)

    if filetype.lower() in ["xlsx", "excel"]:
        rows = []
        for row in data:
            if row[4] == "ROOT":
                continue
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
            rows.append([scaninfo[row[12]][0], lastseen, str(row[4]), str(row[3]),
                         str(row[2]), row[13], datafield])

        if len(ids.split(',')) > 1 or scan_name == "":
            fname = "SpiderFoot.xlsx"
        else:
            fname = _safe_filename(scan_name) + "-SpiderFoot.xlsx"

        return Response(
            build_excel(rows, ["Scan Name", "Updated", "Type", "Module", "Source", "F/P", "Data"], sheet_name_index=2),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{fname}"',
                'Pragma': 'no-cache'
            }
        )

    if filetype.lower() == 'csv':
        fileobj = StringIO()
        parser = csv.writer(fileobj, dialect=dialect)
        parser.writerow(["Scan Name", "Updated", "Type", "Module", "Source", "F/P", "Data"])
        for row in data:
            if row[4] == "ROOT":
                continue
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
            parser.writerow([_csv_safe(scaninfo[row[12]][0]), lastseen, _csv_safe(row[4]), _csv_safe(row[3]),
                             _csv_safe(row[2]), _csv_safe(row[13]), _csv_safe(datafield)])

        if len(ids.split(',')) > 1 or scan_name == "":
            fname = "SpiderFoot.csv"
        else:
            fname = _safe_filename(scan_name) + "-SpiderFoot.csv"

        return Response(
            fileobj.getvalue().encode('utf-8'),
            mimetype='application/csv',
            headers={
                'Content-Disposition': f'attachment; filename="{fname}"',
                'Pragma': 'no-cache'
            }
        )

    return jsonify(["ERROR", "Invalid export filetype."])


# -- Scan search result export -----------------------------------------------

@api_bp.route('/scansearchresultexport', methods=['GET', 'POST'])
def scansearchresultexport():
    """Export search result data as CSV or Excel."""
    id = request.values.get('id', '')
    eventType = request.values.get('eventType', None)
    value = request.values.get('value', None)
    filetype = request.values.get('filetype', 'csv')
    dialect = request.values.get('dialect', 'excel')

    data = search_base(id, eventType, value)

    if not data:
        return Response('', status=204)

    if filetype.lower() in ["xlsx", "excel"]:
        rows = []
        for row in data:
            if row[10] == "ROOT":
                continue
            datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
            rows.append([row[0], str(row[10]), str(row[3]), str(row[2]), row[11], datafield])

        return Response(
            build_excel(rows, ["Updated", "Type", "Module", "Source", "F/P", "Data"], sheet_name_index=1),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': 'attachment; filename=SpiderFoot.xlsx',
                'Pragma': 'no-cache'
            }
        )

    if filetype.lower() == 'csv':
        fileobj = StringIO()
        parser = csv.writer(fileobj, dialect=dialect)
        parser.writerow(["Updated", "Type", "Module", "Source", "F/P", "Data"])
        for row in data:
            if row[10] == "ROOT":
                continue
            datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
            parser.writerow([_csv_safe(row[0]), _csv_safe(row[10]), _csv_safe(row[3]), _csv_safe(row[2]), _csv_safe(row[11]), _csv_safe(datafield)])

        return Response(
            fileobj.getvalue().encode('utf-8'),
            mimetype='application/csv',
            headers={
                'Content-Disposition': 'attachment; filename=SpiderFoot.csv',
                'Pragma': 'no-cache'
            }
        )

    return jsonify(["ERROR", "Invalid export filetype."])


# -- Scan export JSON multi --------------------------------------------------

@api_bp.route('/scanexportjsonmulti', methods=['GET', 'POST'])
def scanexportjsonmulti():
    """Export scan event result data in JSON format for multiple scans."""
    ids = request.values.get('ids', '')

    dbh = get_db()
    scaninfo = list()
    scan_name = ""

    for id in ids.split(','):
        scan = dbh.scanInstanceGet(id)
        if scan is None:
            continue

        scan_name = scan[0]

        for row in dbh.scanResultEvent(id):
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            event_data = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
            source_data = str(row[2])
            source_module = str(row[3])
            event_type = row[4]
            false_positive = row[13]

            if event_type == "ROOT":
                continue

            scaninfo.append({
                "data": event_data,
                "event_type": event_type,
                "module": source_module,
                "source_data": source_data,
                "false_positive": false_positive,
                "last_seen": lastseen,
                "scan_name": scan_name,
                "scan_target": scan[1]
            })

    if len(ids.split(',')) > 1 or scan_name == "":
        fname = "SpiderFoot.json"
    else:
        fname = _safe_filename(scan_name) + "-SpiderFoot.json"

    return Response(
        json.dumps(scaninfo).encode('utf-8'),
        mimetype='application/json; charset=utf-8',
        headers={
            'Content-Disposition': f'attachment; filename="{fname}"',
            'Pragma': 'no-cache'
        }
    )


# -- Scan viz ----------------------------------------------------------------

@api_bp.route('/scanviz', methods=['GET', 'POST'])
def scanviz():
    """Export entities from scan results for visualising."""
    id = request.values.get('id', '')
    gexf = request.values.get('gexf', '0')

    if not id:
        return Response('', status=204)

    dbh = get_db()
    data = dbh.scanResultEvent(id, filterFp=True)
    scan = dbh.scanInstanceGet(id)

    if not scan:
        return Response('', status=204)

    scan_name = scan[0]
    root = scan[1]

    if gexf == "0":
        # buildGraphJson already returns a JSON string; wrapping it in jsonify
        # would double-encode it. Send as a raw application/json response.
        return Response(
            SpiderFootHelpers.buildGraphJson([root], data),
            mimetype='application/json',
        )

    fname = _safe_filename(scan_name) + "-SpiderFoot.gexf"

    return Response(
        SpiderFootHelpers.buildGraphGexf([root], "SpiderFoot Export", data),
        mimetype='application/gexf',
        headers={
            'Content-Disposition': f'attachment; filename="{fname}"',
            'Pragma': 'no-cache'
        }
    )


# -- Scan viz multi ----------------------------------------------------------

@api_bp.route('/scanvizmulti', methods=['GET', 'POST'])
def scanvizmulti():
    """Export entities from multiple scans in GEXF format."""
    ids = request.values.get('ids', '')
    gexf = request.values.get('gexf', '1')

    dbh = get_db()
    data = list()
    roots = list()
    scan_name = ""

    if not ids:
        return Response('', status=204)

    for id in ids.split(','):
        scan = dbh.scanInstanceGet(id)
        if not scan:
            continue
        data = data + dbh.scanResultEvent(id, filterFp=True)
        roots.append(scan[1])
        scan_name = scan[0]

    if not data:
        return Response('', status=204)

    if gexf == "0":
        # Not implemented yet
        return Response('', status=204)

    if len(ids.split(',')) > 1 or scan_name == "":
        fname = "SpiderFoot.gexf"
    else:
        fname = _safe_filename(scan_name) + "-SpiderFoot.gexf"

    return Response(
        SpiderFootHelpers.buildGraphGexf(roots, "SpiderFoot Export", data),
        mimetype='application/gexf',
        headers={
            'Content-Disposition': f'attachment; filename="{fname}"',
            'Pragma': 'no-cache'
        }
    )


# ---------------------------------------------------------------------------
# Scan presets
# ---------------------------------------------------------------------------


@api_bp.route('/presets', methods=['GET'])
def presets_list():
    """List all scan presets (built-in + user)."""
    dbh = get_db()
    return jsonify([serialize_preset(r) for r in dbh.presetList()])


@api_bp.route('/presets/<path:preset_id>', methods=['GET'])
def presets_get(preset_id):
    """Get a single preset by id."""
    dbh = get_db()
    r = dbh.presetGet(preset_id)
    if r is None:
        return jsonify_error('404', f"Preset {preset_id} not found")
    return jsonify(serialize_preset(r))


@api_bp.route('/presets', methods=['POST'])
def presets_create():
    """Create a user preset."""
    dbh = get_db()
    body = request.get_json(silent=True) or {}
    name = (body.get('name') or '').strip()
    description = body.get('description')
    modules = body.get('modules') or []

    if not name or len(name) > 60:
        return jsonify_error('400', "Name must be 1-60 characters")
    if description is not None and not isinstance(description, str):
        return jsonify_error('400', "description must be a string")
    if description is not None and len(description) > 200:
        return jsonify_error('400', "description must be 200 characters or fewer")
    if not isinstance(modules, list):
        return jsonify_error('400', "modules must be a list")

    sf_modules = get_config().get('__modules__') or {}
    valid, invalid = validate_module_names(modules, sf_modules)
    if invalid:
        return jsonify_error('400', f"Unknown modules: {', '.join(invalid)}")

    for existing in dbh.presetList():
        if existing['name'].lower() == name.lower():
            return jsonify_error('400', f"A preset named '{existing['name']}' already exists")

    preset_id = f"user:{uuid.uuid4().hex}"
    now_ms = int(time.time() * 1000)
    try:
        dbh.presetCreate(
            preset_id=preset_id,
            name=name,
            description=description,
            kind='user',
            sort_order=0,
            modules=valid,
            now_ms=now_ms,
        )
    except sqlite3.IntegrityError as e:
        # UNIQUE/CHECK constraint violation — most likely a name collision
        # that slipped past the prior presetList() check (TOCTOU race) or a
        # CHECK violation on `kind`. Always a client-facing data issue.
        return jsonify_error('400', f"Constraint violation: {e}")
    except Exception as e:
        return jsonify_error('500', f"Failed to create preset: {e}")
    return jsonify(serialize_preset(dbh.presetGet(preset_id))), 201


@api_bp.route('/presets/<path:preset_id>', methods=['PATCH'])
def presets_update(preset_id):
    dbh = get_db()
    existing = dbh.presetGet(preset_id)
    if existing is None:
        return jsonify_error('404', f"Preset {preset_id} not found")
    if existing['kind'] == 'builtin':
        return jsonify_error('403', "Built-in presets are read-only")

    body = request.get_json(silent=True) or {}
    name = (body.get('name') or existing['name']).strip()
    description = body.get('description', existing['description'])
    modules = body.get('modules', existing['modules'])

    if not name or len(name) > 60:
        return jsonify_error('400', "Name must be 1-60 characters")
    if description is not None and not isinstance(description, str):
        return jsonify_error('400', "description must be a string")
    if description is not None and len(description) > 200:
        return jsonify_error('400', "description must be 200 characters or fewer")
    if not isinstance(modules, list):
        return jsonify_error('400', "modules must be a list")

    sf_modules = get_config().get('__modules__') or {}
    valid, invalid = validate_module_names(modules, sf_modules)
    if invalid:
        return jsonify_error('400', f"Unknown modules: {', '.join(invalid)}")

    # Name conflict — but allow keeping the same name on this preset
    for other in dbh.presetList():
        if other['id'] == preset_id:
            continue
        if other['name'].lower() == name.lower():
            return jsonify_error('400', f"A preset named '{other['name']}' already exists")

    try:
        dbh.presetUpdate(
            preset_id=preset_id, name=name, description=description,
            modules=valid, now_ms=int(time.time() * 1000),
        )
    except sqlite3.IntegrityError as e:
        return jsonify_error('400', f"Constraint violation: {e}")
    except Exception as e:
        return jsonify_error('500', f"Failed to update preset: {e}")
    return jsonify(serialize_preset(dbh.presetGet(preset_id)))


# IMPORTANT: register the LITERAL /presets/default route BEFORE the
# <path:preset_id> route so Flask matches it first for DELETE.
@api_bp.route('/presets/default', methods=['DELETE'])
def presets_clear_default():
    dbh = get_db()
    dbh.presetClearDefault()
    return jsonify({'status': 'ok'})


@api_bp.route('/presets/<path:preset_id>', methods=['DELETE'])
def presets_delete(preset_id):
    dbh = get_db()
    existing = dbh.presetGet(preset_id)
    if existing is None:
        return jsonify_error('404', f"Preset {preset_id} not found")
    if existing['kind'] == 'builtin':
        return jsonify_error('403', "Built-in presets cannot be deleted")
    dbh.presetDelete(preset_id)
    return jsonify({'status': 'ok'})


@api_bp.route('/presets/<path:preset_id>/default', methods=['POST'])
def presets_set_default(preset_id):
    dbh = get_db()
    if dbh.presetGet(preset_id) is None:
        return jsonify_error('404', f"Preset {preset_id} not found")
    dbh.presetSetDefault(preset_id)
    return jsonify({'status': 'ok'})
